import streamlit as st 
import pandas as pd
import re
from openpyxl import load_workbook
from Sheets_to_Excel import getGoogleSheet 
from Sheets_to_Excel import getFilePath
from Sheets_to_Excel import viewData

st.markdown('# Data Manipulation App')

#st.markdown('''
            ###### \nPlease share your google sheet with the following email address 
#            so that it can be send to Google Cloud for fast loading: 
#            muni-s-test-service-account@symmetric-flare-370800.iam.gserviceaccount.com''')

url = st.text_input('Enter the url of your Google Sheet: ')

# convert Google Sheet to Excel and download to user's downloads folder 
if st.button('Convert Google Sheet to Excel File'):
    getGoogleSheet(url)
    viewData()

# create input fields for the row and column
row = st.number_input('Row:')
column = st.number_input('Column:')

# create an input field for the data to add
data = st.text_input('Data:')

# create a button that, when clicked, will add the data to the specified row and column
if st.button('Add Data'):

    # load the Excel file
    wb = load_workbook(getFilePath())

    # get the active sheet in the Excel file
    ws = wb.active

    # get the cell at the specified row and column
    cell = ws.cell(row=row, column=column)

    # set the value of the cell
    cell.value = data

    # save the Excel file
    wb.save(getFilePath())

    # show a success message
    st.success('Data added successfully!')

    # Display the updated data from the DataFrame
    viewData()

## FIX :: HOW TO ROUND A SPECIFIC COLUMNS VALUES TO 2 DECIMAL POINTS & DISPLAY/SAVE EXCEL FILE PROPERLY